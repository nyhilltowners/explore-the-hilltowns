#!/usr/bin/env python3
"""Atlas build: reads manifest.json, parses each category workbook by its schema,
validates rigorously, and writes site/index.html + site/data.js.

Exit code 1 on any validation failure => GitHub Actions keeps the last good site live.
Only the FIRST sheet of each workbook is read; other sheets are yours for notes/examples.
Hidden-by-policy fields (IDs shown never, Notes shipped never) are enforced here.
"""
import datetime
import html as html_mod
import json
import re
import shutil
import sys
import urllib.request
from pathlib import Path
from urllib.parse import urljoin

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SITE = ROOT / "site"

FAILS: list[str] = []
WARNS: list[str] = []


def fail(msg: str) -> None:
    FAILS.append(msg)


def warn(msg: str) -> None:
    WARNS.append(msg)


def s(v) -> str:
    return "" if v is None else str(v).strip()


def map_headers(header_row, spec, where):
    """spec: list of (key, regex, required). Returns {key: column_index}."""
    idx = {}
    hdrs = [s(h) for h in header_row]
    for key, rx, required in spec:
        for i, h in enumerate(hdrs):
            if h and re.search(rx, h, re.I):
                idx[key] = i
                break
        else:
            if required:
                fail(f"{where}: required column matching /{rx}/ not found "
                     f"(headers seen: {', '.join(h for h in hdrs if h) or 'none'})")
    return idx


def cell(row, idx, key):
    i = idx.get(key)
    if i is None or i >= len(row):
        return ""
    v = row[i]
    return "" if v is None else v


def coord(v, lo, hi, what, where, required):
    if s(v) == "":
        if required:
            fail(f"{where}: {what} is required")
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        fail(f"{where}: {what} '{v}' is not a number")
        return None
    if not (lo <= f <= hi):
        fail(f"{where}: {what} {f} is outside [{lo}, {hi}]")
        return None
    return round(f, 6)


DATE_FMTS = ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%B %d, %Y", "%b %d, %Y", "%d %B %Y")


def parse_date(v, what, where, required):
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    t = s(v)
    if t == "":
        if required:
            fail(f"{where}: {what} is required")
        return None
    for fmt in DATE_FMTS:
        try:
            return datetime.datetime.strptime(t, fmt).date().isoformat()
        except ValueError:
            pass
    fail(f"{where}: {what} '{t}' is not a recognizable date (use YYYY-MM-DD or MM/DD/YYYY)")
    return None


OG_RXS = [
    re.compile(r'<meta[^>]+property=["\']og:image(?::secure_url)?["\'][^>]*content=["\']([^"\']+)["\']', re.I),
    re.compile(r'<meta[^>]+content=["\']([^"\']+)["\'][^>]*property=["\']og:image(?::secure_url)?["\']', re.I),
    re.compile(r'<meta[^>]+name=["\']twitter:image(?::src)?["\'][^>]*content=["\']([^"\']+)["\']', re.I),
    re.compile(r'<meta[^>]+content=["\']([^"\']+)["\'][^>]*name=["\']twitter:image(?::src)?["\']', re.I),
]
_og_cache: dict = {}


def fetch_site_image(url, where):
    """Fetch a site's Open Graph / Twitter preview image. Warnings only."""
    if not re.match(r"^https?://", url, re.I):
        url = "https://" + url
    if url in _og_cache:
        return _og_cache[url]
    img = ""
    try:
        req = urllib.request.Request(
            url, headers={"User-Agent": "Mozilla/5.0 (compatible; AtlasBuild/1.0)"})
        with urllib.request.urlopen(req, timeout=8) as resp:
            page = resp.read(300_000).decode("utf-8", errors="ignore")
            final_url = resp.geturl()
        for rx in OG_RXS:
            m = rx.search(page)
            if m:
                img = urljoin(final_url, html_mod.unescape(m.group(1).strip()))
                break
        if not img:
            warn(f"{where}: no preview image published by {url} — pin will have no photo "
                 f"(you can set one explicitly in the Image column)")
    except Exception as e:
        warn(f"{where}: couldn't reach {url} for a preview image "
             f"({e.__class__.__name__}) — pin will have no photo this build")
    _og_cache[url] = img
    return img


WIKI_FILE_RX = re.compile(
    r"https?://(?:commons|en|www)\.(?:wikimedia|wikipedia)\.org/wiki/File:(.+)$", re.I)


def norm_image(u, where):
    """Accepts: direct URL, a Commons/Wikipedia File: page URL (auto-converted
    to a stable thumbnail URL), or a repo-relative path like images/foo.jpg."""
    u = s(u)
    if not u:
        return ""
    m = WIKI_FILE_RX.match(u)
    if m:
        return ("https://commons.wikimedia.org/wiki/Special:FilePath/"
                + m.group(1) + "?width=800")
    if re.match(r"^https?://", u, re.I):
        if "googleusercontent.com" in u:
            warn(f"{where}: Image is a googleusercontent.com link — these expire "
                 f"without notice; consider saving the file into images/ instead")
        return u
    # repo-relative path
    rel = u.lstrip("./")
    if not (ROOT / rel).exists():
        warn(f"{where}: Image '{u}' is not a URL and no file exists at {rel} — "
             f"it will render blank")
    return rel


def read_rows(path: Path, where: str, sheet=None):
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet:
        if sheet not in wb.sheetnames:
            fail(f"{where}: no sheet named '{sheet}' "
                 f"(found: {', '.join(wb.sheetnames)})")
            wb.close()
            return [], []
        ws = wb[sheet]
    else:
        ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        fail(f"{where}: first sheet is empty")
        return [], []
    return rows[0], rows[1:]


def check_dupe_ids(recs, where):
    seen = {}
    for r in recs:
        rid = r.get("id", "")
        if rid:
            if rid in seen:
                warn(f"{where}: duplicate ID '{rid}' (rows keep working; consider fixing)")
            seen[rid] = True


# ---------------- folklore ----------------

FOLK_SPEC = [
    ("id", r"locked\s*id|^id$", False),
    ("t", r"short\s*title", True),
    ("b", r"tale.*being", False),
    ("n", r"nation|people", False),
    ("p", r"general\s*location", False),
    ("lat", r"^lat", True),
    ("lng", r"^lon|^lng", True),
    ("c", r"certainty", False),
    ("stry", r"story\s*summary|^summary$", False),
    ("tf", r"timeframe", False),
    ("rb", r"recorded\s*by", False),
    ("yr", r"year\s*recorded", False),
    ("pub", r"original\s*publication|publication", False),
    ("src", r"source\s*book|source\s*file", False),
    ("pg", r"page", False),
    ("img", r"^image(\s*url)?$|^photo", False),
    ("cred", r"image\s*credit|^credit$|attribution", False),
    ("web", r"^website$|^url$|^link$", False),
    ("tags", r"tags?|keywords?|themes?", False),
    ("disp", r"^display$|^show$|^visible$", False),
]


def tier_of(cert: str, has_xy: bool) -> str:
    c = cert.lower().strip()
    if not has_xy:
        return "none"
    if c.startswith("fairly"):
        return "fair"
    if c.startswith("precise"):
        return "precise"
    if c.startswith("approx") or c.startswith("two specific"):
        return "approx"
    return "unver"


def parse_folklore(path: Path, label: str, sheet=None):
    where = path.name + (f" [{sheet}]" if sheet else "")
    hdr, rows = read_rows(path, where, sheet)
    idx = map_headers(hdr, FOLK_SPEC, where)
    out = []
    for n, row in enumerate(rows, start=2):
        rid, title = s(cell(row, idx, "id")), s(cell(row, idx, "t"))
        if not rid and not title:
            continue
        if s(cell(row, idx, "disp")).lower() in ("no", "n", "false", "0", "hide", "hidden"):
            continue  # curated out via the Display column
        rw = f"{where} row {n}"
        lat = coord(cell(row, idx, "lat"), -90, 90, "Latitude", rw, required=False)
        lng = coord(cell(row, idx, "lng"), -180, 180, "Longitude", rw, required=False)
        if (lat is None) != (lng is None):
            fail(f"{rw}: has one coordinate but not the other")
        rec = {
            "ty": "folklore", "cat": label, "id": rid,
            "t": title or "(untitled)",
            "b": s(cell(row, idx, "b")), "n": s(cell(row, idx, "n")),
            "p": s(cell(row, idx, "p")),
            "lat": lat, "lng": lng,
            "c": s(cell(row, idx, "c")), "s": s(cell(row, idx, "stry")),
            "tf": s(cell(row, idx, "tf")), "rb": s(cell(row, idx, "rb")),
            "yr": s(cell(row, idx, "yr")), "pub": s(cell(row, idx, "pub")),
            "src": s(cell(row, idx, "src")), "pg": s(cell(row, idx, "pg")),
            "img": norm_image(cell(row, idx, "img"), rw),
            "cred": s(cell(row, idx, "cred")),
            "web": s(cell(row, idx, "web")),
            "tags": [x.strip() for x in re.split(r"[;,]", s(cell(row, idx, "tags"))) if x.strip()],
        }
        # column-drift dedupe, mirrored from the interactive versions
        if rec["b"] and rec["b"] == rec["s"]:
            rec["b"] = ""
        if rec["s"] and rec["s"] == rec["tf"] and rec["b"]:
            rec["s"], rec["b"] = rec["b"], ""
        rec["tier"] = tier_of(rec["c"], lat is not None and lng is not None)
        out.append(rec)
    check_dupe_ids(out, where)
    if not out:
        warn(f"{where}: no data rows found")
    return out


# ---------------- points of interest ----------------

POI_SPEC = [
    ("id", r"^id$|locked\s*id", False),
    ("t", r"^name$", True),
    ("tags", r"tags?|categor|type", False),
    ("addr", r"address", False),
    ("lat", r"^lat", True),
    ("lng", r"^lon|^lng", True),
    ("stry", r"description", False),
    ("mon", r"^mon", False), ("tue", r"^tue", False), ("wed", r"^wed", False),
    ("thu", r"^thu", False), ("fri", r"^fri", False), ("sat", r"^sat", False),
    ("sun", r"^sun", False),
    ("web", r"website|^url$", False),
    ("ph", r"phone", False),
    ("g", r"glyph|icon|symbol", False),
    ("img", r"^image(\s*url)?$|^photo", False),
    ("cred", r"image\s*credit|^credit$|attribution", False),
    ("disp", r"^display$|^show$|^visible$", False),
]


def name_v_or_title(t):
    return t or "(unnamed)"


def parse_poi(path: Path, label: str, sheet=None):
    where = path.name + (f" [{sheet}]" if sheet else "")
    hdr, rows = read_rows(path, where, sheet)
    idx = map_headers(hdr, POI_SPEC, where)
    out = []
    for n, row in enumerate(rows, start=2):
        title = html_mod.unescape(s(cell(row, idx, "t")))
        rid = s(cell(row, idx, "id"))
        if not rid and not title:
            continue
        rw = f"{where} row {n}"
        disp = s(cell(row, idx, "disp")).lower()
        if disp in ("no", "n", "false", "0", "hide", "hidden"):
            continue  # curated out via the Display column
        if not title:
            fail(f"{rw}: Name is required")
        lat = coord(cell(row, idx, "lat"), -90, 90, "Latitude", rw, required=False)
        lng = coord(cell(row, idx, "lng"), -180, 180, "Longitude", rw, required=False)
        if lat is None or lng is None:
            # not yet locatable — keep it in the catalog list, just off the map
            warn(f"{rw}: {name_v_or_title(title)} has no coordinates yet — "
                 f"listed but not mapped")
        web = s(cell(row, idx, "web"))
        if web and not re.match(r"^(https?://)?[\w.-]+\.[a-z]{2,}", web, re.I):
            warn(f"{rw}: Website '{web}' doesn't look like a URL")
        hrs = {d: s(cell(row, idx, d)) for d in
               ("mon", "tue", "wed", "thu", "fri", "sat", "sun")}
        # markers (the base "Points of Interest" layer) don't have hours by design
        raw_tags = s(cell(row, idx, "tags"))
        tags = [x.strip() for x in re.split(r"[;,]", raw_tags) if x.strip()]
        out.append({
            "ty": "poi", "cat": label, "id": rid, "t": title or "(unnamed)",
            "tags": tags, "addr": s(cell(row, idx, "addr")),
            "lat": lat, "lng": lng, "s": s(cell(row, idx, "stry")),
            "hrs": hrs, "web": web, "ph": s(cell(row, idx, "ph")),
            "g": s(cell(row, idx, "g")),
            "tier": "exact" if (lat is not None and lng is not None) else "none",
            "img": norm_image(cell(row, idx, "img"), rw),
            "cred": s(cell(row, idx, "cred")),
        })
        if not out[-1]["img"] and web and label != "Points of Interest":
            out[-1]["img"] = fetch_site_image(web, rw)
    check_dupe_ids(out, where)
    return out


# ---------------- events ----------------

_WD = {"mon": 0, "tue": 1, "wed": 2, "thu": 3, "fri": 4, "sat": 5, "sun": 6}
_ORD = {"1st": 1, "2nd": 2, "3rd": 3, "4th": 4, "5th": 5, "last": 0}

_ORDWORD = {"first": "1st", "second": "2nd", "third": "3rd",
            "fourth": "4th", "fifth": "5th", "last": "last"}


def _normalize_ordinal_recurrence(text):
    """Turn natural standing-schedule phrasing into the 'Recurring — …' grammar
    that next_occurrence() understands. Recognizes forms like:
        '1st & 3rd Tuesdays, 6:00 PM'
        '2nd Wednesdays'
        'first and third Tuesday of the month'
        '2nd & 4th Wednesdays'
    Returns a normalized 'Recurring — 1st week: Tue; 3rd week: Tue' string, or ''
    if no ordinal-weekday pattern is present. This lets schedules be written the
    plain way a person would, while the date math stays in one place."""
    if not text:
        return ""
    t = text.lower()
    # word ordinals -> numeric ('first' -> '1st') so one grammar handles both
    for word, num in _ORDWORD.items():
        t = re.sub(r"\b" + word + r"\b", num, t)
    # find the weekday this clause is about (first weekday token present)
    wdm = re.search(r"\b(mon|tue|wed|thu|fri|sat|sun)", t)
    if not wdm:
        return ""
    wd = wdm.group(1)
    # collect every ordinal that appears before the weekday token
    head = t[:wdm.start()]
    ords = re.findall(r"\b(1st|2nd|3rd|4th|5th|last)\b", head)
    if not ords:
        return ""
    wd_title = wd.capitalize()
    segs = ["{} week: {}".format(o, wd_title) for o in ords]
    return "Recurring — " + "; ".join(segs)


def _next_weekday(frm, wd):
    """First date >= frm falling on weekday wd (0=Mon..6=Sun)."""
    delta = (wd - frm.weekday()) % 7
    return frm + datetime.timedelta(days=delta)


def _nth_weekday_of_month(year, month, wd, nth):
    """The nth (1..5, or 0='last') weekday wd in a given month, or None."""
    first = datetime.date(year, month, 1)
    first_wd = _next_weekday(first, wd)
    if nth == 0:  # last occurrence
        d = first_wd
        while (d + datetime.timedelta(days=7)).month == month:
            d += datetime.timedelta(days=7)
        return d
    d = first_wd + datetime.timedelta(days=7 * (nth - 1))
    return d if d.month == month else None


def next_occurrence(notes, start, end, today, time_field=""):
    """Given a 'Recurring — …' Notes string, return the soonest date >= today
    that the event actually happens (respecting its start/end bounds), or None
    if the notes carry no parseable recurrence. Handles:
      'every week: Thursday'            (one or many days)
      '2nd week: Tuesday'               (nth weekday of month)
      '1st week of month: Wednesday; 3rd week: Wednesday'  (compound)
    If the Notes carry no explicit 'Recurring —' clause, natural standing-schedule
    phrasing in the Notes or Time field (e.g. '1st & 3rd Tuesdays, 6:00 PM') is
    normalized into the same grammar, so plainly-written schedules still surface.
    Dates are ISO strings; today/start/end are ISO strings or ''.
    """
    m = re.search(r"Recurring\s*[\u2014-]\s*([^|]+)", notes or "")
    if not m:
        # no explicit recurrence clause — try to read natural ordinal phrasing
        # from the Notes, then the Time field ('1st & 3rd Tuesdays, 6:00 PM')
        normalized = (_normalize_ordinal_recurrence(notes)
                      or _normalize_ordinal_recurrence(time_field))
        if not normalized:
            return None
        m = re.search(r"Recurring\s*[\u2014-]\s*([^|]+)", normalized)
        if not m:
            return None
    frag = m.group(1).strip()

    def iso2d(x):
        try:
            return datetime.date.fromisoformat(x[:10])
        except Exception:
            return None

    tdy = iso2d(today)
    st = iso2d(start) if start else None
    en = iso2d(end) if end else None
    lo = max(st, tdy) if st else tdy
    if lo is None:
        return None

    def within(d):
        return d is not None and d >= lo and (en is None or d <= en)

    cands = []

    wk = re.search(r"every week\s*:\s*(.+)", frag, re.I)
    if wk:
        for name in re.findall(r"Mon|Tue|Wed|Thu|Fri|Sat|Sun", wk.group(1)):
            d = _next_weekday(lo, _WD[name.lower()])
            if within(d):
                cands.append(d)

    for seg in re.split(r";", frag):
        mm = re.search(r"(1st|2nd|3rd|4th|5th|last)\s*week(?:\s*of\s*month)?\s*:\s*"
                       r"(Mon|Tue|Wed|Thu|Fri|Sat|Sun)", seg, re.I)
        if not mm:
            continue
        nth = _ORD[mm.group(1).lower()]
        wd = _WD[mm.group(2).lower()]
        probe = datetime.date(lo.year, lo.month, 1)
        for _ in range(4):
            d = _nth_weekday_of_month(probe.year, probe.month, wd, nth)
            if within(d) and d >= lo:
                cands.append(d)
                break
            probe = (probe.replace(day=28) + datetime.timedelta(days=7)).replace(day=1)

    return min(cands).isoformat() if cands else None


EVT_SPEC = [
    ("id", r"^id$", False),
    ("t", r"event\s*name|^name$|^title", True),
    ("ven", r"venue|location\s*name", False),
    ("addr", r"address", False),
    ("lat", r"^lat", True),
    ("lng", r"^lon|^lng", True),
    ("stry", r"description", False),
    ("tags", r"tags?|keywords?|themes?", False),
    ("d1", r"start\s*date|^date$", True),
    ("d2", r"end\s*date", False),
    ("tm", r"^time", False),
    ("web", r"website|^url$", False),
    ("g", r"glyph|icon|symbol", False),
    ("img", r"^image(\s*url)?$|^photo", False),
    ("cred", r"image\s*credit|^credit$|attribution", False),
    ("notes", r"^notes?$", False),
    ("disp", r"^display$|^show$|^visible$", False),
]



# ---- Event glyph auto-fill -------------------------------------------------
# Applied at build time whenever an event's Glyph cell is blank, so newly
# scraped events get sensible emojis without manual passes. Ordered: the
# first matching rule wins; titles are consulted before venues so an event's
# own subject beats its host's identity. Hand-set cells are never overridden
# (except the lossless 🎬→🍿 normalization below).
GLYPH_NORMALIZE = {"🎬": "🍿"}
GLYPH_RULES = [
 (r"\bfood truck(s)?\b", "\U0001F69A"),
 (r"\bkids?\b|\bchild(ren)?(\'s)?\b|\btoddlers?\b|\byouth\b|\bteens?\b|\bstory ?time\b|\bfor ages? \d|\blittle ones\b|\bknee high\b", "🐤"),
 (r"\bmovies?\b|\bfilm(s|ing)?\b|\bscreening\b|\bdrive[- ]?in\b|\bcinema\b", "🍿"),
 (r"\bart gallery\b|\bgallery (opening|show|night|tour)\b|\bexhibit(ion)?\b|\bopening reception\b|\bcurator tour\b", "🖼️"),
 (r"\btheat(re|er)\b|\bstage play\b|\ba play\b|\bmusical\b|\bcabaret\b|\bimprov\b|\bcomedy (show|night|festival)\b", "🎟️"),
 (r"\bchoir\b", "👯"),
 (r"\bjazz\b", "🎶"),
 (r"\bopen[- ]?mics?\b|\bkaraoke\b", "🎤"),
 (r"\bconcert\b|\bsymphony\b|\bquartet\b|\bchamber music\b|\brecital\b|\bsingalong\b|\bbluegrass\b|\blive music\b", "🎵"),
 (r"\bperformances?\b|\bintermission\b|\bplayers present\b|\bshakespeare\b", "🎭"),
 (r"\bbutterfl(y|ies)\b|\blupine fest\b|\bmonarch\b(?! hill)", "🦋"),
 (r"\bturtles?\b", "🐢"),
 (r"\btennis\b", "🎾"),
 (r"\bhorse(s|back)?\b|\bequestrian\b|\bpon(y|ies)\b|\btrail(s)? ?rides?\b|\bmustang\b|\bunbridled\b", "🐴"),
 (r"\bbird(s|ing| watch| banding| walk)?\b|\beagle walk\b|\baudubon\b", "🐦"),
 (r"\bdye(s|ing)? workshop\b|\b(plant|natural) dyes?\b|\bdyeing\b", "🎨"),
 (r"\bacupuncture\b", "😌"),
 (r"\bpostpartum\b|\bnew parents?\b|\bbabywearing\b|\bla leche\b", "🍼"),
 (r"\bforest (walk|bath(e|ing)|stewardship|tour|school)\b", "🌲"),
 (r"\bnature bus\b", "🚌"),
 (r"\bastrolog(y|ical|er)\b|\bhoroscopes?\b|\btarot\b|\breikk?i\b|\bpsychics?\b", "🔮"),
 (r"\bbee ?keep(er|ers|ing)?\b|\bbeekep\w*\b", "🍯"),
 (r"\bnative trees?\b", "🌲"),
 (r"\bnative plants?\b", "🌿"),
 (r"\bzba\b|\btown (hall|board) meeting\b|\bbudget town hall\b", "🇺🇸"),
 (r"(?<!baseball )\bbats?\b(?! mitzvah)(?!man)|\bbat (walk|night|watch|count)\b", "🦇"),
 (r"\bhik(e|es|ing)\b|\btrail (run|walk|preview|day)\b|\btrail ?blaz(e|ing|er|ers)\b|\bramble\b|\bmeander\b|\bnature walk\b|\bwalks?\b|\bwalking\b", "🥾"),
 (r"\bart(s)?\b|\bpaint(ing)?\b|\bdraw(ing)?\b|\bsketch\b|\bwatercolor\b|\bprintmaking\b|\bscreenprint(ing)?\b|\bpottery\b|\bceramics\b|\bsculpture\b|\bcreative circles?\b", "🎨"),
 (r"\bwrit(ing|ers?)\b|\bscreenwrit\w*\b|\bmemoir\b|\bjournaling\b", "🖋️"),
 (r"\bbook (club|fair|launch|sale|signing)\b|\blibrary\b|\bauthor\b|\bpoetry\b|\bpoem\b|\breading\b", "📚"),
 (r"\btrivia\b|\bquiz\b", "🧠"),
 (r"\bpuzzles?\b|\bjigsaw\b|\bcrossword\b", "🧩"),
 (r"\bmah ?jongg?\b", "🀄"),
 (r"\bboard games?\b|\bgame night\b|\bgames? meetup\b|\btabletop\b|\bwarmachine\b|\bwargam(e|es|ing)\b|\bmagic:? the gathering\b|\borganized play\b|\bcasual play\b|\bclocktower\b|\bcrokinole\b|\bdart(s| league)\b|\bbingo\b|\bchess\b", "🎲"),
 (r"\bbrewery\b|\bbrewing\b|\bbeer\b|\bcask\b|\btap ?takeover\b|\bcider\b|\bmeadworks\b|\bmead\b", "🍺"),
 (r"\bcocktail\b|\bhappy hour\b|\bwine (tasting|dinner|pairing)\b|\bsip\b", "🍸"),
 (r"\bsound ?bath\b|\bsound healing\b|\bsound(s)? for healing\b", "🎐"),
 (r"\byoga\b|\bmeditat(ion|e)\b|\bbreath ?work\b", "🧘"),
 (r"\bcoworking\b|\btech meetup\b|\bcod(e|ing)\b", "💼"),
 (r"\bfarmers market\b|\bfarm stand\b|\bplant (sale|swap|exchange)\b|\bseed (swap|sowing|library)\b|\bgarden(ing)?\b", "🌱"),
 (r"\bforag(e|ing)\b|\bmushrooms?\b|\bfungi\b|\bmycelium\b|\bmycolog(y|ical|ist)\b", "🍄"),
 (r"\bstitch(ing|ery)?\b|\bup-?stitch\b|\bquilt(s|ing)?\b|\bcross-?stitch\b", "🧵"),
 (r"\bknit(ting)?\b|\bpurl\b|\bcrochet\b|\bsew(ing)?\b|\bcraft\b|\bweav(e|ing)\b|\bembroidery\b|\bmending\b", "🧶"),
 (r"\b5k\b|\brun club\b|\btrot\b|\bfun run\b", "🏃"),
 (r"\bkayak\b|\bpaddle\b|\bcanoe\b|\bsail\b", "🛶"),
 (r"\bgala\b|\bpotluck\b|\bpart(y|ies)\b|\bcelebration\b|\banniversary\b|\bbirthday\b", "🎉"),
 (r"\bvolunteer\b|\bclean ?up\b|\btrash crawl\b|\bfood pantry\b|\bfood drive\b|\bdonation\b|\bfundraiser\b|\bbenefit\b", "🤝"),
 (r"\bfestival\b|\bfair\b|\bfest\b|\bcarnival\b", "🎪"),
 (r"\bhistor(y|ic|ical)\b|\bmuseum\b|\bheritage\b|\brevolution\b|\bcivil war\b", "🏛️"),
]
def auto_glyph(title, venue, cell_glyph, is_online=False):
    # Food trucks always show the truck, overriding any celled food glyph
    # (🍕/🍔/etc.) so the category reads consistently at a glance on the map.
    if re.search(r"\bfood truck(s)?\b", (title or "").lower()):
        return "\U0001F69A"  # 🚚
    g = (cell_glyph or "").strip()
    if g:
        # 💻 changed meaning: it now marks online events. A celled 💻 on an
        # in-person event predates that change and meant coworking → 💼.
        if g == "💻" and not is_online:
            return "💼"
        return GLYPH_NORMALIZE.get(g, g)
    t = (title or "").lower()
    for pat, gg in GLYPH_RULES:          # pass 1: the event's own title
        if re.search(pat, t):
            return gg
    tv = t + " " + (venue or "").lower() # pass 2: fall back to venue context
    for pat, gg in GLYPH_RULES:
        if re.search(pat, tv):
            return gg
    return "💻" if is_online else ""

def parse_events(path: Path, label: str, sheet=None):
    where = path.name + (f" [{sheet}]" if sheet else "")
    hdr, rows = read_rows(path, where, sheet)
    idx = map_headers(hdr, EVT_SPEC, where)
    today = datetime.date.today().isoformat()
    out = []
    skipped_past = [0]
    for n, row in enumerate(rows, start=2):
        title = html_mod.unescape(s(cell(row, idx, "t")))
        rid = s(cell(row, idx, "id"))
        if not rid and not title:
            continue
        rw = f"{where} row {n}"
        if s(cell(row, idx, "disp")).lower() in ("no", "n", "false", "0", "hide", "hidden"):
            continue  # curated out via the Display column (blank = shown)
        if not title:
            fail(f"{rw}: Event Name is required")
        ven_txt = s(cell(row, idx, "ven")).lower()
        addr_txt = s(cell(row, idx, "addr")).lower()
        is_online = bool(re.search(r"\bonline\b|\bvirtual\b|\bzoom\b(?!\s*flume)|\blivestream\b",
                                   ven_txt + " " + addr_txt))
        lat = coord(cell(row, idx, "lat"), -90, 90, "Latitude", rw,
                    required=False)
        lng = coord(cell(row, idx, "lng"), -180, 180, "Longitude", rw,
                    required=False)
        # Food trucks are roaming vendors. When their location is only a town
        # (or blank) — no street address — they have no fixed spot to pin, so we
        # keep them in the listings/search but off the map. A real street address
        # (e.g. hosted at a brewery) keeps them mapped as normal.
        _addr_raw = s(cell(row, idx, "addr")).strip()
        _has_street = bool(re.match(r"^\d+\s+\S", _addr_raw))
        if re.search(r"\bfood truck(s)?\b", (title or "").lower()) and not _has_street:
            if lat is not None or lng is not None:
                warn(f"{rw}: {title} is a roaming food truck with no street address — listed/searchable but not mapped")
            lat = None
            lng = None
        if not is_online and (lat is None or lng is None):
            # upcoming but not yet locatable: keep it in the listings, off the map
            warn(f"{rw}: {title} has no coordinates yet — listed but not mapped")
        d1 = parse_date(cell(row, idx, "d1"), "Start Date", rw, required=True)
        d2 = parse_date(cell(row, idx, "d2"), "End Date", rw, required=False)
        if d1 and d2 and d2 < d1:
            fail(f"{rw}: End Date {d2} is before Start Date {d1}")
        notes = s(cell(row, idx, "notes"))
        nocc = next_occurrence(notes, d1, d2, today,
                               time_field=s(cell(row, idx, "tm")))  # ISO string or None
        end = d2 or d1
        if end and end < today and not nocc:
            skipped_past[0] += 1
            continue  # non-recurring past events are dropped from the build
        if nocc and end and end < today:
            # a recurring series whose bare End Date is stale but still recurs:
            # its real horizon is the next occurrence, so it stays in.
            pass
        out.append({
            "ty": "events", "cat": label, "id": rid, "t": title or "(unnamed)",
            "ven": html_mod.unescape(s(cell(row, idx, "ven"))), "addr": s(cell(row, idx, "addr")),
            "tags": [x.strip() for x in re.split(r"[;,]", s(cell(row, idx, "tags"))) if x.strip()],
            "lat": lat, "lng": lng, "s": html_mod.unescape(s(cell(row, idx, "stry"))),
            "d1": d1, "d2": d2 or "", "tm": s(cell(row, idx, "tm")),
            "web": s(cell(row, idx, "web")), "g": auto_glyph(title, s(cell(row, idx, "ven")), s(cell(row, idx, "g")), is_online),
            "online": is_online, "nextOcc": nocc or "",
            "tier": "exact" if (lat is not None and lng is not None) else "none",
            "img": norm_image(cell(row, idx, "img"), rw),
            "cred": s(cell(row, idx, "cred")),
        })
        ev_web = s(cell(row, idx, "web"))
        if not out[-1]["img"] and ev_web:
            out[-1]["img"] = fetch_site_image(ev_web, rw)
    check_dupe_ids(out, where)
    if skipped_past[0]:
        print(f"    ({skipped_past[0]} past events skipped)")
    return out


PARSERS = {"folklore": parse_folklore, "poi": parse_poi, "events": parse_events}


def resolve_workbook(path: Path, label: str) -> Path:
    """Exact filename wins. Otherwise accept exactly one versioned variant
    (points_of_interest_v2.xlsx, points_of_interest (3).xlsx, etc.).
    Multiple candidates = ambiguous -> loud stop so the wrong data never ships."""
    variants = sorted(p for p in path.parent.glob(path.stem + "*" + path.suffix)
                      if p.name != path.name and not p.name.startswith("~$"))
    if path.exists():
        if variants:
            warn(f"{label}: using {path.name}, but versioned copies also exist "
                 f"({', '.join(v.name for v in variants)}) — delete extras or "
                 f"the wrong data may be published")
        return path
    if len(variants) == 1:
        warn(f"{label}: {path.name} not found; using {variants[0].name} instead")
        return variants[0]
    if len(variants) > 1:
        fail(f"{label}: {path.name} not found and multiple candidates exist "
             f"({', '.join(v.name for v in variants)}) — keep exactly one")
    return path


def backfill_event_coords_from_poi(all_records):
    """Fill in coordinates for events that lack them by reusing the verified
    coordinates of a matching Point of Interest. Matching is deliberately strict:
    an event borrows a POI's location only when its venue name equals the POI name,
    or its street address matches the POI's street address. This never invents a
    coordinate — it only reuses one already vetted in the POI sheet — and it never
    overrides coordinates an event already has. Loose/substring matching is avoided
    on purpose so we don't, say, pin a 'Sleepy Hollow' reading to a farm that merely
    shares a word."""
    def norm(v):
        return re.sub(r"[^a-z0-9]+", " ", str(v or "").lower()).strip()

    def street_key(addr):
        m = re.match(r"\s*(\d+)\s+([a-z0-9]+)", str(addr or "").lower())
        return (m.group(1), m.group(2)) if m else None

    pois = [r for r in all_records
            if r.get("ty") == "poi" and r.get("lat") is not None]
    by_name, by_addr = {}, {}
    for p in pois:
        n = norm(p.get("t"))
        if n and n not in by_name:
            by_name[n] = p
        k = street_key(p.get("addr"))
        if k and k not in by_addr:
            by_addr[k] = p

    filled = 0
    for e in all_records:
        if e.get("ty") != "events" or e.get("lat") is not None:
            continue
        if e.get("online"):
            continue
        hit = by_name.get(norm(e.get("ven")))
        if not hit:
            k = street_key(e.get("addr"))
            if k:
                hit = by_addr.get(k)
        if hit:
            e["lat"], e["lng"] = hit["lat"], hit["lng"]
            e["tier"] = "exact"
            filled += 1
    if filled:
        print(f"  event→POI coord backfill: filled {filled} event(s) from matching POIs")


def main() -> int:
    manifest = json.loads((ROOT / "manifest.json").read_text(encoding="utf-8"))
    all_records = []
    categories = []
    for cat in manifest["categories"]:
        label, schema = cat["label"], cat["schema"]
        wb_path = resolve_workbook(ROOT / cat["workbook"], label)
        categories.append({"key": cat["key"], "label": label, "schema": schema,
                           "color": cat.get("color", ""),
                           "default_on": cat.get("default_on", True),
                           "glyph": cat.get("glyph", "")})
        if schema not in PARSERS:
            fail(f"manifest: unknown schema '{schema}' for '{label}'")
            continue
        if not wb_path.exists():
            warn(f"manifest: workbook {cat['workbook']} not found — '{label}' will be empty")
            continue
        recs = PARSERS[schema](wb_path, label, cat.get("sheet"))
        dg = cat.get("glyph", "")
        if dg:
            for r in recs:
                if not r.get("g"):
                    r["g"] = dg
        all_records.extend(recs)
        print(f"  {label}: {len(recs)} records")

    backfill_event_coords_from_poi(all_records)

    print()
    for w in WARNS:
        print(f"WARNING: {w}")
    if FAILS:
        print()
        for f_ in FAILS:
            print(f"ERROR:   {f_}")
        print(f"\nBUILD FAILED — {len(FAILS)} error(s). "
              f"The live site keeps its last good version until these are fixed.")
        return 1

    SITE.mkdir(exist_ok=True)
    # Remove stale generated HTML so renamed/removed pages don't linger locally.
    for old in SITE.glob("*.html"):
        old.unlink()
    payload = {
        "generated": datetime.datetime.now(datetime.timezone.utc)
                     .strftime("%Y-%m-%d %H:%M UTC"),
        "categories": categories,
        "records": all_records,
    }
    js = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    js = js.replace("</", "<\\/")
    (SITE / "data.js").write_text("window.ATLAS_DATA = " + js + ";\n",
                                  encoding="utf-8")
    template = ROOT / "index.template.html"
    if not template.exists() and (ROOT / "index.html").exists():
        template = ROOT / "index.html"  # tolerate the template being renamed
    if not template.exists():
        print("ERROR: no index.template.html (or index.html) found at repo root")
        return 1
    shutil.copyfile(template, SITE / "index.html")
    # Standalone About page (hand-authored, not templated).
    if (ROOT / "about.html").exists():
        shutil.copyfile(ROOT / "about.html", SITE / "about.html")
    # Standalone Event Calendar page (reads site/data.js at runtime).
    if (ROOT / "calendar.html").exists():
        shutil.copyfile(ROOT / "calendar.html", SITE / "calendar.html")
    if (ROOT / "images").exists():
        shutil.copytree(ROOT / "images", SITE / "images", dirs_exist_ok=True)
    if (ROOT / "fonts").exists():
        shutil.copytree(ROOT / "fonts", SITE / "fonts", dirs_exist_ok=True)
    mapped = sum(1 for r in all_records if r.get("lat") is not None)
    print(f"\nBUILD OK — {len(all_records)} records ({mapped} mappable), "
          f"{len(WARNS)} warning(s). Output in site/.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
